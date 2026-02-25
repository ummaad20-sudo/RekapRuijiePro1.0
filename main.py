import os

# =============================
# ANDROID PERMISSION RUNTIME
# =============================
try:
    from android.permissions import request_permissions, Permission
    request_permissions([
        Permission.READ_EXTERNAL_STORAGE,
        Permission.WRITE_EXTERNAL_STORAGE,
        Permission.MANAGE_EXTERNAL_STORAGE,
    ])
except Exception:
    pass

# =============================
# IMPORT
# =============================
from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.filechooser import FileChooserListView
from kivy.uix.button import Button
from kivy.uix.label import Label
from kivy.uix.scrollview import ScrollView
from kivy.uix.popup import Popup
from openpyxl import load_workbook
from collections import defaultdict
from datetime import datetime


class RekapApp(App):

    def build(self):
        root = BoxLayout(orientation='vertical', padding=10, spacing=10)

        # ===== TITLE =====
        title = Label(
            text="Rekap Data Ruijie Pro",
            size_hint=(1, 0.1)
        )
        root.add_widget(title)

        # ===== HASIL (SCROLL) =====
        self.result_label = Label(
            text="Silakan pilih file Excel...",
            size_hint_y=None,
            valign='top'
        )
        self.result_label.bind(
            texture_size=self.result_label.setter('size')
        )

        scroll = ScrollView(size_hint=(1, 0.8))
        scroll.add_widget(self.result_label)
        root.add_widget(scroll)

        # ===== BUTTON DI BAWAH =====
        btn = Button(
            text="Pilih File Excel",
            size_hint=(1, 0.1)
        )
        btn.bind(on_press=self.buka_file)
        root.add_widget(btn)

        return root

    # =============================
    # FILE PICKER
    # =============================
    def buka_file(self, instance):
        try:
            start_path = "/storage/emulated/0/Download"
            if not os.path.exists(start_path):
                start_path = "/sdcard/Download"

            chooser = FileChooserListView(
                path=start_path,
                filters=["*.xlsx"]
            )

            popup = Popup(
                title="Pilih File Excel",
                content=chooser,
                size_hint=(0.95, 0.95)
            )

            chooser.bind(
                on_submit=lambda x, sel, touch:
                self.proses_file(sel, popup)
            )

            popup.open()

        except Exception as e:
            self.result_label.text = f"Gagal membuka file picker:\n{e}"

    # =============================
    # PROSES FILE
    # =============================
    def proses_file(self, selection, popup):
        if not selection:
            self.result_label.text = "Tidak ada file dipilih."
            return

        file_path = selection[0]
        popup.dismiss()

        try:
            wb = load_workbook(file_path, data_only=True)
            sheet = wb.active
        except Exception as e:
            self.result_label.text = f"Gagal membuka file:\n{e}"
            return

        rekap_detail = defaultdict(lambda: {"jumlah": 0, "total": 0})
        rekap_tanggal = defaultdict(int)

        header = [cell.value for cell in sheet[1]]

        try:
            kolom_grup = header.index("Grup pengguna")
            kolom_harga = header.index("Harga")
            kolom_tanggal = header.index("Diaktifkan di")
        except ValueError:
            self.result_label.text = (
                "Header tidak sesuai!\n"
                "Pastikan ada:\n"
                "- Grup pengguna\n"
                "- Harga\n"
                "- Diaktifkan di"
            )
            return

        # =============================
        # LOOP DATA
        # =============================
        for row in sheet.iter_rows(min_row=2, values_only=True):
            grup = row[kolom_grup]
            harga = row[kolom_harga]
            tanggal_full = row[kolom_tanggal]

            if grup and harga and tanggal_full:

                if isinstance(tanggal_full, datetime):
                    tanggal = tanggal_full.strftime("%Y/%m/%d")
                else:
                    tanggal = str(tanggal_full).split(" ")[0]

                try:
                    harga_int = int(float(str(harga).replace(",", "").replace(".", "")))
                except Exception:
                    continue

                key = (tanggal, grup)
                rekap_detail[key]["jumlah"] += 1
                rekap_detail[key]["total"] += harga_int
                rekap_tanggal[tanggal] += harga_int

        # =============================
        # FORMAT HASIL
        # =============================
        hasil = "===== HASIL REKAP =====\n\n"
        tanggal_terakhir = None

        for (tanggal, grup), data in sorted(rekap_detail.items()):

            if tanggal_terakhir and tanggal != tanggal_terakhir:
                hasil += f">>> TOTAL {tanggal_terakhir} : Rp {rekap_tanggal[tanggal_terakhir]:,}\n"
                hasil += "-----------------------------\n"

            if tanggal != tanggal_terakhir:
                hasil += f"\nTanggal : {tanggal}\n"

            hasil += f"  Grup   : {grup}\n"
            hasil += f"  Jumlah : {data['jumlah']}\n"
            hasil += f"  Total  : Rp {data['total']:,}\n\n"

            tanggal_terakhir = tanggal

        if tanggal_terakhir:
            hasil += f">>> TOTAL {tanggal_terakhir} : Rp {rekap_tanggal[tanggal_terakhir]:,}\n"

        self.result_label.text = hasil


if __name__ == "__main__":
    RekapApp().run()
